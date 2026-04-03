import calendar
import csv
import io
import os
import sqlite3
import unicodedata
from collections import defaultdict
from contextlib import closing
from datetime import date, datetime
from functools import wraps
from zoneinfo import ZoneInfo

from flask import (
    Flask,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    session,
    abort,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, "timecard.db")
COMPANY_NAME = "株式会社まごころグループ"
DEFAULT_ADMIN_PASSWORD = "admin1234"
TOKYO_TZ = ZoneInfo("Asia/Tokyo")

app = Flask(__name__)
app.config["SECRET_KEY"] = "change-this-secret-key-v30"
app.config["SESSION_COOKIE_NAME"] = "magokoro_timecard_v30"


ALL_STORES = [
    "からだﾗﾎﾞ駅前院",
    "からだﾗﾎﾞ桜台院",
    "ﾗｺﾝｼｪﾙ青葉台店",
    "鍼灸ﾗｺﾝｼｪﾙ青葉台店",
    "からだﾗﾎﾞ溝の口院",
    "ﾗｺﾝｼｪﾙ三軒茶屋",
    "からだﾗﾎﾞ駒沢大学駅院",
    "狛江駅前整骨院",
    "からだﾗﾎﾞ新百合ヶ丘北口院",
    "からだﾗﾎﾞ溝の口分院",
    "からだﾗﾎﾞたまﾌﾟﾗｰｻﾞ院",
    "鍼灸ﾗｺﾝｼｪﾙたまﾌﾟﾗｰｻﾞ店",
    "ﾗｺﾝｼｪﾙ溝の口店",
    "鍼灸ﾗｺﾝｼｪﾙ代々木上原店",
    "ラ・コンシェル町田店",
    "からだラボ整骨院　センター北院",
    "からだラボ整骨院　マプレ院",
    "からだﾗﾎﾞ武蔵小杉院",
    "からだﾗﾎﾞ用賀院",
    "からだﾗﾎﾞ向ヶ丘遊園院",
]


# ---------- DB helpers ----------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        ensure_runtime_schema(g.db)
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def ensure_runtime_schema(db):
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS user_store_permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            store_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, store_id),
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (store_id) REFERENCES stores (id)
        )
        """
    )
    db.commit()


def normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "")
    text = text.replace("　", " ")
    return " ".join(text.split())


def detect_brand_and_store(display_name: str) -> tuple[str, str]:
    normalized = normalize_text(display_name)

    if normalized.startswith("鍼灸ラコンシェル"):
        return "鍼灸ラコンシェル", normalized.replace("鍼灸ラコンシェル", "", 1).strip()
    if normalized.startswith("ラコンシェル"):
        return "ラコンシェル", normalized.replace("ラコンシェル", "", 1).strip()
    if normalized.startswith("ラ・コンシェル"):
        return "ラコンシェル", normalized.replace("ラ・コンシェル", "", 1).strip()
    if normalized.startswith("からだラボ整骨院"):
        return "からだラボ整骨院", normalized.replace("からだラボ整骨院", "", 1).strip()
    if normalized.startswith("からだラボ"):
        return "からだラボ整骨院", normalized.replace("からだラボ", "", 1).strip()

    if "整骨院" in normalized:
        return "からだラボ整骨院", normalized

    return "からだラボ整骨院", normalized


def sync_master_stores(db):
    existing = {normalize_text(row["display_name"]): row["id"] for row in db.execute("SELECT id, display_name FROM stores").fetchall()}
    admin_hash = generate_password_hash(DEFAULT_ADMIN_PASSWORD)
    first_store = db.execute("SELECT id FROM stores ORDER BY id LIMIT 1").fetchone()
    admin_user_id = db.execute(
        "INSERT INTO users (store_id, username, password_hash, full_name, role, is_active) VALUES (?, ?, ?, ?, ?, 1)",
        (first_store["id"], "admin", admin_hash, "全体管理者", "admin"),
    ).lastrowid

    sample_staff = [
        ("kawahara", "河原", "からだラボ整骨院", "溝の口院", ["からだﾗﾎﾞ溝の口院", "からだﾗﾎﾞ溝の口分院"]),
        ("kaneko", "金子", "からだラボ整骨院", "武蔵小杉院", []),
        ("iwamura", "岩村", "ラコンシェル", "町田店", ["ラ・コンシェル町田店"]),
        ("hanada", "花田", "鍼灸ラコンシェル", "代々木上原店", ["鍼灸ﾗｺﾝｼｪﾙ代々木上原店"]),
    ]
    for username, full_name, brand_name, store_name, managed_display_names in sample_staff:
        store = db.execute(
            "SELECT id FROM stores WHERE brand_name = ? AND store_name = ? LIMIT 1",
            (brand_name, store_name),
        ).fetchone()
        if store:
            user_id = db.execute(
                "INSERT INTO users (store_id, username, password_hash, full_name, role, is_active) VALUES (?, ?, ?, ?, ?, 1)",
                (store["id"], username, generate_password_hash("staff1234"), full_name, "staff"),
            ).lastrowid
            for display_name in managed_display_names:
                managed_store = db.execute(
                    "SELECT id FROM stores WHERE display_name = ? LIMIT 1",
                    (normalize_text(display_name),),
                ).fetchone()
                if managed_store:
                    db.execute(
                        "INSERT OR IGNORE INTO user_store_permissions (user_id, store_id) VALUES (?, ?)",
                        (user_id, managed_store["id"]),
                    )

    db.commit()


# ---------- auth / permissions ----------
def login_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if session.get("user_id") is None:
            return redirect(url_for("login"))
        return view(**kwargs)

    return wrapped_view


def admin_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if session.get("role") != "admin":
            flash("全体管理者のみ利用できます。", "error")
            return redirect(url_for("dashboard"))
        return view(**kwargs)

    return wrapped_view


def store_manager_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if session.get("role") == "admin" or getattr(g, "can_manage_stores", False):
            return view(**kwargs)
        flash("この機能は責任者以上のみ利用できます。", "error")
        return redirect(url_for("dashboard"))

    return wrapped_view


def get_primary_store_id(user_row) -> int | None:
    return user_row["store_id"] if user_row and user_row["store_id"] else None


def get_managed_store_ids(user_id: int, primary_store_id: int | None = None) -> list[int]:
    db = get_db()
    rows = db.execute(
        "SELECT store_id FROM user_store_permissions WHERE user_id = ? ORDER BY store_id",
        (user_id,),
    ).fetchall()
    ids = {row["store_id"] for row in rows}
    if primary_store_id and rows:
        ids.add(primary_store_id)
    return sorted(ids)


def get_accessible_store_ids(user_row) -> list[int] | None:
    if not user_row:
        return []
    if user_row["role"] == "admin":
        return None
    managed_ids = get_managed_store_ids(user_row["id"], user_row["store_id"])
    return managed_ids


def can_manage_stores(user_row) -> bool:
    if not user_row:
        return False
    if user_row["role"] == "admin":
        return True
    return len(get_managed_store_ids(user_row["id"], user_row["store_id"])) > 0


def user_can_access_store(user_row, store_id: int) -> bool:
    ids = get_accessible_store_ids(user_row)
    if ids is None:
        return True
    return store_id in ids


@app.before_request
def load_logged_in_user():
    if request.endpoint == "init_db_route":
        g.user = None
        g.can_manage_stores = False
        g.accessible_store_ids = []
        return

    if not database_ready():
        session.clear()
        g.user = None
        g.can_manage_stores = False
        g.accessible_store_ids = []
        return

    user_id = session.get("user_id")
    if user_id is None:
        g.user = None
        g.can_manage_stores = False
        g.accessible_store_ids = []
    else:
        g.user = get_db().execute(
            """
            SELECT u.*, s.brand_name, s.store_name, s.display_name, s.company_name
            FROM users u
            LEFT JOIN stores s ON u.store_id = s.id
            WHERE u.id = ?
            """,
            (user_id,),
        ).fetchone()
        if g.user is None:
            session.clear()
            g.can_manage_stores = False
            g.accessible_store_ids = []
        else:
            g.can_manage_stores = can_manage_stores(g.user)
            g.accessible_store_ids = get_accessible_store_ids(g.user) or []


# ---------- date / time ----------
def tokyo_now():
    return datetime.now(TOKYO_TZ)


def tokyo_today():
    return tokyo_now().date()


def parse_dt(value: str | None):
    if not value:
        return None
    dt = datetime.fromisoformat(value)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=TOKYO_TZ)
    return dt.astimezone(TOKYO_TZ)


def format_dt(dt_str):
    dt = parse_dt(dt_str)
    if not dt:
        return None
    return dt.strftime("%Y-%m-%d %H:%M")


def format_time_only(dt_str):
    dt = parse_dt(dt_str)
    if not dt:
        return ""
    return dt.strftime("%H:%M")


def jp_weekday(date_str: str) -> str:
    d = datetime.fromisoformat(date_str).date()
    return ["月", "火", "水", "木", "金", "土", "日"][d.weekday()]


def minutes_to_hhmm(minutes):
    if minutes in (None, ""):
        return ""
    minutes = int(minutes)
    return f"{minutes // 60}:{minutes % 60:02d}"


def minutes_to_hours_text(minutes):
    if minutes is None:
        return "-"
    hours = int(minutes) // 60
    mins = int(minutes) % 60
    return f"{hours}時間 {mins:02d}分"


def calculate_work_minutes(clock_in, clock_out, break_minutes):
    start = parse_dt(clock_in)
    end = parse_dt(clock_out)
    if not start or not end:
        return None
    total = int((end - start).total_seconds() // 60) - int(break_minutes or 0)
    return max(total, 0)


app.jinja_env.filters["format_dt"] = format_dt
app.jinja_env.filters["minutes_to_hours_text"] = minutes_to_hours_text
app.jinja_env.filters["minutes_to_hhmm"] = minutes_to_hhmm


# ---------- business logic ----------
def get_today_attendance(user_id: int):
    return get_db().execute(
        "SELECT * FROM attendance WHERE user_id = ? AND work_date = ?",
        (user_id, tokyo_today().isoformat()),
    ).fetchone()


def get_allowed_store_filter(requested_store_id: str | None):
    db = get_db()
    if g.user["role"] == "admin":
        stores = db.execute("SELECT * FROM stores ORDER BY brand_name, store_name").fetchall()
        if requested_store_id:
            return requested_store_id, stores, None
        return None, stores, None

    accessible_ids = g.accessible_store_ids
    if not accessible_ids:
        return None, [], []

    placeholders = ",".join(["?"] * len(accessible_ids))
    stores = db.execute(
        f"SELECT * FROM stores WHERE id IN ({placeholders}) ORDER BY brand_name, store_name",
        accessible_ids,
    ).fetchall()

    if requested_store_id:
        try:
            store_id_int = int(requested_store_id)
        except ValueError:
            flash("店舗指定が不正です。", "error")
            return None, stores, accessible_ids
        if store_id_int not in accessible_ids:
            flash("担当外の店舗は閲覧できません。", "error")
            return None, stores, accessible_ids
        return requested_store_id, stores, [store_id_int]

    return None, stores, accessible_ids


def get_permission_display_map() -> dict[int, str]:
    db = get_db()
    rows = db.execute(
        """
        SELECT usp.user_id, GROUP_CONCAT(s.display_name, ' / ') AS managed_names
        FROM user_store_permissions usp
        JOIN stores s ON usp.store_id = s.id
        GROUP BY usp.user_id
        """
    ).fetchall()
    return {row["user_id"]: row["managed_names"] for row in rows}


def set_user_store_permissions(user_id: int, store_ids: list[str]):
    db = get_db()
    db.execute("DELETE FROM user_store_permissions WHERE user_id = ?", (user_id,))
    for store_id in store_ids:
        db.execute(
            "INSERT OR IGNORE INTO user_store_permissions (user_id, store_id) VALUES (?, ?)",
            (user_id, int(store_id)),
        )
    db.commit()


def resolve_store_id(db, value: str | None):
    normalized = normalize_text(value or "")
    if not normalized:
        return None

    row = db.execute("SELECT id FROM stores WHERE display_name = ? LIMIT 1", (normalized,)).fetchone()
    if row:
        return row["id"]

    row = db.execute("SELECT id FROM stores WHERE store_name = ? LIMIT 1", (normalized,)).fetchone()
    if row:
        return row["id"]

    # exact normalized compare against company-entered names
    rows = db.execute("SELECT id, display_name, store_name FROM stores").fetchall()
    for row in rows:
        if normalize_text(row["display_name"]) == normalized or normalize_text(row["store_name"]) == normalized:
            return row["id"]
    return None


def create_user_record(db, store_id: int, username: str, full_name: str, password: str, role: str, permission_store_ids: list[int]):
    user_id = db.execute(
        "INSERT INTO users (store_id, username, password_hash, full_name, role, is_active) VALUES (?, ?, ?, ?, ?, 1)",
        (store_id, username, generate_password_hash(password), full_name, role),
    ).lastrowid
    if role == "staff":
        for perm_store_id in permission_store_ids:
            db.execute(
                "INSERT OR IGNORE INTO user_store_permissions (user_id, store_id) VALUES (?, ?)",
                (user_id, int(perm_store_id)),
            )
    return user_id


def process_bulk_users(db, raw_text: str):
    raw_text = (raw_text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not raw_text:
        return 0, []

    created = 0
    errors = []
    stream = io.StringIO(raw_text)
    reader = csv.reader(stream)

    for idx, row in enumerate(reader, start=1):
        row = [normalize_text(col) for col in row]
        if not any(row):
            continue
        if len(row) < 4:
            errors.append(f"{idx}行目: 項目数が足りません。『氏名,ユーザーID,初期PW,主所属店舗,権限,担当店舗1|担当店舗2』で入力してください。")
            continue

        full_name, username, password, primary_store_name = row[:4]
        role = row[4] if len(row) >= 5 and row[4] else "staff"
        permissions_raw = row[5] if len(row) >= 6 else ""

        if role not in ("admin", "staff"):
            errors.append(f"{idx}行目: 権限は admin または staff のみ使えます。")
            continue

        store_id = resolve_store_id(db, primary_store_name)
        if not store_id:
            errors.append(f"{idx}行目: 主所属店舗『{primary_store_name}』が見つかりません。")
            continue

        permission_store_ids = []
        if permissions_raw:
            for name in [normalize_text(x) for x in permissions_raw.split("|") if normalize_text(x)]:
                perm_id = resolve_store_id(db, name)
                if not perm_id:
                    errors.append(f"{idx}行目: 担当店舗『{name}』が見つかりません。")
                    permission_store_ids = None
                    break
                permission_store_ids.append(perm_id)
        if permission_store_ids is None:
            continue

        try:
            create_user_record(db, store_id, username, full_name, password, role, permission_store_ids)
            created += 1
        except sqlite3.IntegrityError:
            errors.append(f"{idx}行目: ユーザーID『{username}』が重複しています。")

    db.commit()
    return created, errors


def query_attendance_records(month: str, scoped_store_ids: list[int] | None, selected_store_id: str | None):
    db = get_db()
    query = """
        SELECT a.*, u.full_name, u.username, s.brand_name, s.store_name, s.display_name
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        JOIN stores s ON a.store_id = s.id
        WHERE substr(a.work_date, 1, 7) = ?
    """
    params: list[object] = [month]

    if selected_store_id:
        query += " AND a.store_id = ?"
        params.append(int(selected_store_id))
    elif scoped_store_ids is not None:
        if not scoped_store_ids:
            return []
        placeholders = ",".join(["?"] * len(scoped_store_ids))
        query += f" AND a.store_id IN ({placeholders})"
        params.extend(scoped_store_ids)

    query += " ORDER BY a.work_date DESC, s.brand_name, s.store_name, u.full_name"
    return db.execute(query, params).fetchall()


def query_summary(month: str, scoped_store_ids: list[int] | None, selected_store_id: str | None):
    db = get_db()
    query = """
        SELECT s.id, s.brand_name, s.store_name, s.display_name,
               COUNT(DISTINCT a.user_id) AS staff_count,
               COALESCE(SUM(a.work_minutes), 0) AS total_minutes
        FROM stores s
        LEFT JOIN attendance a ON a.store_id = s.id AND substr(a.work_date, 1, 7) = ?
        WHERE 1=1
    """
    params: list[object] = [month]

    if selected_store_id:
        query += " AND s.id = ?"
        params.append(int(selected_store_id))
    elif scoped_store_ids is not None:
        if not scoped_store_ids:
            return []
        placeholders = ",".join(["?"] * len(scoped_store_ids))
        query += f" AND s.id IN ({placeholders})"
        params.extend(scoped_store_ids)

    query += " GROUP BY s.id ORDER BY s.brand_name, s.store_name"
    return db.execute(query, params).fetchall()


# ---------- routes ----------
@app.route("/init-db")
def init_db_route():
    session.clear()

    existing_db = g.pop("db", None)
    if existing_db is not None:
        existing_db.close()

    if not os.path.exists(DATABASE) or not database_ready():
        init_db()
        return f"DB initialized. Login with admin / {DEFAULT_ADMIN_PASSWORD}"

    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    ensure_runtime_schema(db)
    sync_master_stores(db)
    seed_data(db)
    db.close()
    return "DB already initialized. Existing users and attendance were kept."


@app.route("/", methods=["GET", "POST"])
def login():
    if not database_ready():
        return redirect(url_for("init_db_route"))

    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        db = get_db()
        user = db.execute(
            """
            SELECT u.*, s.brand_name, s.store_name, s.display_name, s.company_name
            FROM users u
            LEFT JOIN stores s ON u.store_id = s.id
            WHERE u.username = ? AND u.is_active = 1
            """,
            (username,),
        ).fetchone()
        error = None
        if user is None or not check_password_hash(user["password_hash"], password):
            error = "ログイン情報が正しくありません。"

        if error is None:
            session.clear()
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            return redirect(url_for("dashboard"))

        flash(error, "error")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    today_attendance = get_today_attendance(g.user["id"])
    recent_logs = get_db().execute(
        """
        SELECT * FROM attendance
        WHERE user_id = ?
        ORDER BY work_date DESC
        LIMIT 10
        """,
        (g.user["id"],),
    ).fetchall()
    return render_template(
        "dashboard.html",
        today_attendance=today_attendance,
        recent_logs=recent_logs,
    )


@app.route("/punch/<action>", methods=["POST"])
@login_required
def punch(action):
    db = get_db()
    record = get_today_attendance(g.user["id"])
    now = tokyo_now().replace(second=0, microsecond=0).isoformat()

    if action == "clock_in":
        if record and record["clock_in"]:
            flash("本日の出勤打刻は完了しています。", "error")
        else:
            if record:
                db.execute(
                    "UPDATE attendance SET clock_in = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (now, record["id"]),
                )
            else:
                db.execute(
                    "INSERT INTO attendance (user_id, store_id, work_date, clock_in) VALUES (?, ?, ?, ?)",
                    (g.user["id"], g.user["store_id"], tokyo_today().isoformat(), now),
                )
            db.commit()
            flash("出勤打刻しました。", "success")

    elif action == "break_start":
        if not record or not record["clock_in"]:
            flash("先に出勤打刻をしてください。", "error")
        elif record["break_start"] and not record["break_end"]:
            flash("すでに休憩中です。", "error")
        else:
            db.execute(
                "UPDATE attendance SET break_start = ?, break_end = NULL, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (now, record["id"]),
            )
            db.commit()
            flash("休憩開始を記録しました。", "success")

    elif action == "break_end":
        if not record or not record["break_start"]:
            flash("休憩開始が未打刻です。", "error")
        elif record["break_end"]:
            flash("休憩終了はすでに記録済みです。", "error")
        else:
            start = parse_dt(record["break_start"])
            end = parse_dt(now)
            delta_minutes = int((end - start).total_seconds() // 60)
            total_break = (record["break_minutes"] or 0) + max(delta_minutes, 0)
            db.execute(
                "UPDATE attendance SET break_end = ?, break_minutes = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (now, total_break, record["id"]),
            )
            db.commit()
            flash("休憩終了を記録しました。", "success")

    elif action == "clock_out":
        if not record or not record["clock_in"]:
            flash("先に出勤打刻をしてください。", "error")
        elif record["clock_out"]:
            flash("本日の退勤打刻は完了しています。", "error")
        else:
            break_minutes = record["break_minutes"] or 0
            work_minutes = calculate_work_minutes(record["clock_in"], now, break_minutes)
            db.execute(
                "UPDATE attendance SET clock_out = ?, work_minutes = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (now, work_minutes, record["id"]),
            )
            db.commit()
            flash("退勤打刻しました。", "success")
    else:
        flash("不正な操作です。", "error")

    return redirect(url_for("dashboard"))


@app.route("/admin")
@login_required
@store_manager_required
def admin_panel():
    month = request.args.get("month") or tokyo_today().strftime("%Y-%m")
    requested_store_id = request.args.get("store_id") or None
    selected_store_id, stores, scoped_store_ids = get_allowed_store_filter(requested_store_id)

    records = query_attendance_records(month, scoped_store_ids, selected_store_id)
    summary = query_summary(month, scoped_store_ids, selected_store_id)

    return render_template(
        "admin.html",
        records=records,
        stores=stores,
        summary=summary,
        selected_month=month,
        selected_store_id=selected_store_id or "",
        can_view_all=(g.user["role"] == "admin"),
    )


@app.route("/admin/users", methods=["GET", "POST"])
@login_required
@admin_required
def manage_users():
    db = get_db()
    if request.method == "POST":
        bulk_text = request.form.get("bulk_text", "").strip()
        if bulk_text:
            created, errors = process_bulk_users(db, bulk_text)
            if created:
                flash(f"一括登録で {created} 名を追加しました。", "success")
            for error in errors[:10]:
                flash(error, "error")
            if errors and len(errors) > 10:
                flash(f"ほか {len(errors) - 10} 件のエラーがあります。入力を見直してください。", "error")
            return redirect(url_for("manage_users"))

        store_id = request.form["store_id"]
        username = request.form["username"].strip()
        full_name = request.form["full_name"].strip()
        password = request.form["password"]
        role = request.form["role"]
        permission_store_ids = [int(x) for x in request.form.getlist("permission_store_ids")]
        try:
            create_user_record(db, int(store_id), username, full_name, password, role, permission_store_ids)
            db.commit()
            flash("スタッフを追加しました。", "success")
        except sqlite3.IntegrityError:
            flash("ユーザー名が重複しています。", "error")
        return redirect(url_for("manage_users"))

    users = db.execute(
        """
        SELECT u.*, s.brand_name, s.store_name, s.display_name
        FROM users u
        LEFT JOIN stores s ON u.store_id = s.id
        ORDER BY s.brand_name, s.store_name, u.full_name
        """
    ).fetchall()
    stores = db.execute("SELECT * FROM stores ORDER BY brand_name, store_name").fetchall()
    permission_map = get_permission_display_map()
    return render_template("users.html", users=users, stores=stores, permission_map=permission_map)


@app.route("/admin/users/<int:user_id>/permissions", methods=["GET", "POST"])
@login_required
@admin_required
def edit_user_permissions(user_id: int):
    db = get_db()
    user = db.execute(
        """
        SELECT u.*, s.display_name
        FROM users u
        LEFT JOIN stores s ON u.store_id = s.id
        WHERE u.id = ?
        """,
        (user_id,),
    ).fetchone()
    if not user:
        flash("対象スタッフが見つかりません。", "error")
        return redirect(url_for("manage_users"))

    stores = db.execute("SELECT * FROM stores ORDER BY brand_name, store_name").fetchall()
    selected_store_ids = {
        row["store_id"]
        for row in db.execute("SELECT store_id FROM user_store_permissions WHERE user_id = ?", (user_id,)).fetchall()
    }

    if request.method == "POST":
        permission_store_ids = request.form.getlist("permission_store_ids")
        set_user_store_permissions(user_id, permission_store_ids)
        flash("担当店舗を更新しました。", "success")
        return redirect(url_for("manage_users"))

    return render_template(
        "user_permissions.html",
        user=user,
        stores=stores,
        selected_store_ids=selected_store_ids,
    )


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete_user(user_id: int):
    db = get_db()
    user = db.execute(
        "SELECT id, full_name, username, role FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()

    if user is None:
        flash("対象スタッフが見つかりません。", "error")
        return redirect(url_for("manage_users"))

    if user["id"] == session.get("user_id"):
        flash("ログイン中の自分自身は削除できません。", "error")
        return redirect(url_for("manage_users"))

    db.execute("DELETE FROM user_store_permissions WHERE user_id = ?", (user_id,))
    db.execute("DELETE FROM attendance WHERE user_id = ?", (user_id,))
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    flash(f"{user['full_name']}（{user['username']}）を削除しました。", "success")
    return redirect(url_for("manage_users"))


def attendance_rows_for_export(month: str, scoped_store_ids: list[int] | None, selected_store_id: str | None):
    return query_attendance_records(month, scoped_store_ids, selected_store_id)


@app.route("/admin/export")
@login_required
@store_manager_required
def export_csv():
    month = request.args.get("month") or tokyo_today().strftime("%Y-%m")
    requested_store_id = request.args.get("store_id") or None
    selected_store_id, _stores, scoped_store_ids = get_allowed_store_filter(requested_store_id)
    rows = attendance_rows_for_export(month, scoped_store_ids, selected_store_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "スタッフ名", "ブランド", "店舗", "日", "曜日", "始業時刻", "終業時刻", "休憩時間", "実働時間", "備考"
    ])
    for row in rows:
        day_date = datetime.fromisoformat(row["work_date"]).date()
        writer.writerow([
            row["full_name"],
            row["brand_name"],
            row["display_name"],
            f"{day_date.day}日",
            jp_weekday(row["work_date"]),
            format_time_only(row["clock_in"]),
            format_time_only(row["clock_out"]),
            minutes_to_hhmm(row["break_minutes"] or 0),
            minutes_to_hhmm(row["work_minutes"]),
            row["note"] or "",
        ])

    mem = io.BytesIO()
    mem.write(output.getvalue().encode("utf-8-sig"))
    mem.seek(0)

    filename = f"timecard_{month}.csv"
    return send_file(mem, as_attachment=True, download_name=filename, mimetype="text/csv")


def build_workbook(month: str, rows):
    year, month_num = map(int, month.split("-"))
    days_in_month = calendar.monthrange(year, month_num)[1]

    grouped = defaultdict(dict)
    profile = {}
    for row in rows:
        grouped[row["full_name"]][row["work_date"]] = row
        profile[row["full_name"]] = row

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    yellow_fill = PatternFill("solid", fgColor="FFF7B2")
    gray_fill = PatternFill("solid", fgColor="EDEDED")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for idx, (full_name, row_map) in enumerate(sorted(grouped.items()), start=1):
        sheet_name = full_name[:28] or f"sheet{idx}"
        ws = wb.create_sheet(title=sheet_name)
        info = profile[full_name]

        ws["A1"] = f"{year}年{month_num}月勤務時間計算表"
        ws["A1"].font = Font(size=18, bold=True)
        ws.merge_cells("A1:G1")

        ws["B2"] = "氏 名"
        ws["C2"] = full_name
        ws["F2"] = "店舗"
        ws["G2"] = info["display_name"]
        for cell in ["B2", "C2", "F2", "G2"]:
            ws[cell].border = border
            ws[cell].alignment = center
        ws["B2"].fill = header_fill
        ws["F2"].fill = header_fill
        ws["C2"].fill = yellow_fill
        ws["G2"].fill = yellow_fill

        headers = ["日", "曜日", "始業時刻", "終業時刻", "休憩時間", "実働時間", "備考"]
        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border

        total_work_minutes = 0
        worked_days = 0
        weekday_map = ["月", "火", "水", "木", "金", "土", "日"]
        for day in range(1, days_in_month + 1):
            row_idx = 4 + day
            work_date = date(year, month_num, day).isoformat()
            record = row_map.get(work_date)
            weekday = weekday_map[date(year, month_num, day).weekday()]

            values = [
                f"{day}日",
                weekday,
                format_time_only(record["clock_in"]) if record else "",
                format_time_only(record["clock_out"]) if record else "",
                minutes_to_hhmm(record["break_minutes"] or 0) if record else "",
                minutes_to_hhmm(record["work_minutes"]) if record else "",
                record["note"] if record and record["note"] else "",
            ]

            if record and record["work_minutes"]:
                total_work_minutes += int(record["work_minutes"])
                worked_days += 1

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = center if col_idx < 7 else Alignment(vertical="center")
                if col_idx in (3, 4):
                    cell.fill = yellow_fill
                elif col_idx in (5, 6):
                    cell.fill = gray_fill

        summary_row = 6 + days_in_month
        ws.cell(row=summary_row, column=1, value="実働日数").border = border
        ws.cell(row=summary_row, column=2, value=worked_days).border = border
        ws.cell(row=summary_row, column=4, value=f"{month_num}月分").alignment = center
        ws.cell(row=summary_row, column=5, value="実働時間").border = border
        ws.cell(row=summary_row, column=6, value=minutes_to_hours_text(total_work_minutes)).border = border

        for col in "ABCDEFG":
            ws.column_dimensions[col].width = 14
        ws.column_dimensions["G"].width = 24

    if not grouped:
        ws = wb.create_sheet(title="勤務表")
        ws["A1"] = f"{month} の勤務データがありません"

    return wb


@app.route("/admin/export-excel")
@login_required
@store_manager_required
def export_excel():
    month = request.args.get("month") or tokyo_today().strftime("%Y-%m")
    requested_store_id = request.args.get("store_id") or None
    selected_store_id, _stores, scoped_store_ids = get_allowed_store_filter(requested_store_id)
    rows = attendance_rows_for_export(month, scoped_store_ids, selected_store_id)

    wb = build_workbook(month, rows)
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    filename = f"timecard_sheet_{month}.xlsx"
    return send_file(
        mem,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/edit/<int:attendance_id>", methods=["GET", "POST"])
@login_required
@store_manager_required
def edit_attendance(attendance_id):
    db = get_db()
    record = db.execute(
        """
        SELECT a.*, u.full_name, s.brand_name, s.store_name, s.display_name
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        JOIN stores s ON a.store_id = s.id
        WHERE a.id = ?
        """,
        (attendance_id,),
    ).fetchone()
    if not record:
        flash("対象データが見つかりません。", "error")
        return redirect(url_for("admin_panel"))

    if not user_can_access_store(g.user, record["store_id"]):
        flash("担当外の店舗データは編集できません。", "error")
        return redirect(url_for("admin_panel"))

    if request.method == "POST":
        clock_in = request.form.get("clock_in") or None
        break_start = request.form.get("break_start") or None
        break_end = request.form.get("break_end") or None
        clock_out = request.form.get("clock_out") or None
        note = request.form.get("note") or None
        break_minutes = int(request.form.get("break_minutes") or 0)
        work_minutes = calculate_work_minutes(clock_in, clock_out, break_minutes)
        db.execute(
            """
            UPDATE attendance
            SET clock_in = ?, break_start = ?, break_end = ?, clock_out = ?,
                break_minutes = ?, work_minutes = ?, note = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (clock_in, break_start, break_end, clock_out, break_minutes, work_minutes, note, attendance_id),
        )
        db.commit()
        flash("打刻を更新しました。", "success")
        return redirect(url_for("admin_panel"))

    return render_template("edit_attendance.html", record=record)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
